# ======================================
# click_scheduler.py (using pyautogui & logging)
# Usage: python click_scheduler.py hh:mm:ss:msms [x y]
# Schedules a global mouse click at the specified clock time (24h),
# then logs the click details into results.xlsx in the same folder.
# Requires: pip install pyautogui openpyxl
# On macOS: enable Accessibility permissions for Terminal/VSCode.
# ======================================

import sys
import time
import os
from datetime import datetime, timedelta
import pyautogui
from openpyxl import load_workbook, Workbook


def parse_time(timestr: str) -> datetime:
    """
    Parse hh:mm:ss:msms into next-occurrence datetime.
    """
    try:
        hh, mm, ss, ms = [int(p) for p in timestr.split(':')]
        now = datetime.now()
        target = now.replace(hour=hh, minute=mm, second=ss, microsecond=ms*1000)
        if target <= now:
            target += timedelta(days=1)
        return target
    except Exception:
        raise ValueError("Time format must be hh:mm:ss:msms, e.g. 14:05:30:250")


def log_result(date_str: str, time_str: str, x: int, y: int):
    """
    Append click record to results.xlsx in script directory.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, 'results.xlsx')

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Time', 'x', 'y'])

    ws.append([date_str, time_str, x, y])
    wb.save(file_path)
    return file_path


def main():
    usage = (
        "Usage: python click_scheduler.py hh:mm:ss:msms [x y]\n"
        "  hh:mm:ss:msms -> target clock time in 24h format\n"
        "  x y           -> optional click coordinates (default: 532 457)\n"
    )
    if len(sys.argv) == 1 or sys.argv[1] in ('-h','--help'):
        print(usage)
        sys.exit(0)
    if len(sys.argv) not in (2,4):
        print(usage)
        sys.exit(1)

    try:
        target_time = parse_time(sys.argv[1])
    except ValueError as e:
        print(e)
        print(usage)
        sys.exit(1)

    if len(sys.argv) == 4:
        try:
            x = int(sys.argv[2]); y = int(sys.argv[3])
        except ValueError:
            print("Coordinates must be integers.")
            print(usage)
            sys.exit(1)
    else:
        x, y = 532, 457

    now = datetime.now()
    delay = (target_time - now).total_seconds()

    print(usage)
    print(f"Scheduled click at {target_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
    print(f"Position: ({x},{y}); waiting {delay:.3f}s...")

    time.sleep(delay)
    pyautogui.moveTo(x, y)
    pyautogui.click()

    # Record the actual click time
    click_time = datetime.now()
    date_str = click_time.strftime('%Y-%m-%d')
    time_str = click_time.strftime('%H:%M:%S.%f')[:-3]
    print(f"Clicked at ({x},{y}) at {date_str} {time_str}")

    # Log to Excel
    file_path = log_result(date_str, time_str, x, y)
    print(f"Logged result to: {file_path}")


if __name__ == '__main__':
    main()
